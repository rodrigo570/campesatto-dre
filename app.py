from datetime import datetime
import io
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go

# ════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="DRE Campesatto — Processador",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stMetric"] { background:#fff; border-radius:10px; padding:16px; border:1px solid #E8ECF0; }
.stButton>button { background:#1A2744; color:#fff; border:none; border-radius:8px; font-weight:500; width:100%; }
.stButton>button:hover { background:#2E4A7A; color:#fff; }
.stDownloadButton>button { background:#059669; color:#fff; border:none; border-radius:8px; font-weight:500; width:100%; }
.stDownloadButton>button:hover { background:#047857; color:#fff; }
div[data-testid="stFileUploader"] { border:2px dashed #CBD5E1; border-radius:10px; padding:8px; }
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════
# HELPERS FORMATAÇÃO
# ════════════════════════════════════════════════════════════
def fmtM(v):
    if v is None: return "—"
    m = abs(float(v)) / 1_000_000
    s = f"{m:,.1f}M".replace(",","X").replace(".",",").replace("X",".")
    return f"R$ {s}" if float(v) >= 0 else f"(R$ {s})"

def fmtK(v):
    if v is None: return "—"
    k = abs(float(v)) / 1_000
    s = f"{k:,.1f}k".replace(",","X").replace(".",",").replace("X",".")
    return f"R$ {s}" if float(v) >= 0 else f"(R$ {s})"

def pct_str(v, base):
    if not base: return "—"
    return f"{v/base*100:.1f}%".replace(".",",")

# ════════════════════════════════════════════════════════════
# EXCEL — helpers de estilo
# ════════════════════════════════════════════════════════════
C = dict(
    DARK='1A2744', MED='2E4A7A', LIGHT='BBDEFB', WHITE='FFFFFF',
    GRAY='F5F7FA', GRAY2='ECEFF1',
    ORIG='FFF3E0', ORIG_FG='E65100',
    ADJ='EDE7F6',  ADJ_FG='4A148C',
    CORR='E8F5E9', CORR_FG='1B5E20',
    NOTE='FFF8E1', NOTE_FG='856404',
    TEAL='004D40',
)

def fl(k):
    h = C.get(k, k).lstrip('#')
    return PatternFill('solid', fgColor='FF'+h if len(h)==6 else h)

def ft(bold=False, sz=9, color='333333', italic=False):
    c = C.get(color, color).lstrip('#')
    if len(c) == 6: c = 'FF' + c
    return Font(bold=bold, size=sz, color=c, name='Arial', italic=italic)

def al(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bdr(thick=False):
    s = 'medium' if thick else 'thin'
    return Border(
        bottom=Side(border_style=s, color='FF8AA6C4'),
        left=Side(border_style='thin', color='FFDDDDDD'),
    )

# ════════════════════════════════════════════════════════════
# DETECÇÃO DO TIPO DE ARQUIVO
# ════════════════════════════════════════════════════════════
def detectar_tipo(wb):
    """
    mensal      → 1 aba 'Planilha1', 3 colunas (Código, Conta, Valor mês)
    obras       → múltiplas abas com nomes de obras
    consolidada → 1 aba principal, muitas colunas de meses
    """
    EXCLUIR = {'obras em andamento','fluxo de caixa','rental','dashboard','fs energia-cnp'}
    abas_dados = [s for s in wb.sheetnames if s.lower() not in EXCLUIR]

    if len(abas_dados) >= 4:
        return 'obras'

    ws = wb[abas_dados[0]]
    max_col = ws.max_column

    # Planilha1 com poucas colunas = mensal
    if abas_dados[0].lower() in ('planilha1','sheet1','plan1') and max_col <= 5:
        return 'mensal'

    # Muitas colunas = consolidada anual com meses
    if max_col > 10:
        return 'consolidada'

    return 'mensal'

# ════════════════════════════════════════════════════════════
# PROCESSAMENTO — DRE MENSAL (Planilha1, 3 colunas)
# ════════════════════════════════════════════════════════════
def processar_mensal(wb, nome, ajustes, da_pct):
    aba = next((s for s in wb.sheetnames
                if s.lower() in ('planilha1','sheet1','plan1')), wb.sheetnames[0])
    ws = wb[aba]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Detectar período (header row 1, col C)
    periodo = str(rows[0][2]).strip() if rows and rows[0][2] else 'Mês'

    # Mapear código → valor
    km = {}
    for row in rows[1:]:
        code = ' '.join(str(row[0]).strip().split()) if row[0] else ''
        if not code: continue
        try:
            v = float(row[2])
            km[code] = v
        except (TypeError, ValueError):
            pass

    def gv(code): return km.get(code, 0.0)

    # Valores
    rec_bruta  = gv('01.01.01')
    deducoes   = gv('01.01.02')
    irpj       = gv('01.01.02.01.06')
    csll       = gv('01.01.02.01.05')
    outras_rec = gv('01.01.03.01.04')   # empréstimos lançados como receita
    custos_dir = gv('02.01.01')
    custos_ind = gv('02.01.02')
    capex      = gv('02.01.02.05.04')
    desp_adm   = gv('03.01')
    desp_trib  = gv('03.02')
    desp_fin   = gv('03.03')
    juros      = gv('03.03.01.03')
    amort      = gv('03.03.01.07')
    nop        = gv('04')
    resultado  = gv('05')

    # Ajustes
    aj1 = -capex          if ajustes[0] else 0  # remove CAPEX do custo
    aj2 = -(irpj + csll)  if ajustes[1] else 0  # remove IRPJ+CSLL das deduções
    aj3 = -amort          if ajustes[2] else 0  # remove amort das desp fin
    aj4 = abs(rec_bruta) * da_pct if ajustes[3] else 0  # D&A
    aj5 = -outras_rec     if outras_rec > 0 else 0       # remove empréstimos de receita

    res_corr = resultado + aj1 + aj2 + aj3 + aj4 + aj5
    # EBITDA = resultado corrigido (conta 05 corrigido) - desp fin corrigida + D&A
    desp_fin_corr = desp_fin - amort if ajustes[2] else desp_fin
    ebitda_corr = res_corr - desp_fin_corr + aj4 - nop

    return {
        'tipo': 'mensal', 'arquivo': nome, 'periodo': periodo,
        'rec_bruta': rec_bruta, 'deducoes': deducoes,
        'custos_dir': custos_dir, 'custos_ind': custos_ind,
        'capex': capex, 'desp_adm': desp_adm, 'desp_trib': desp_trib,
        'desp_fin': desp_fin, 'juros': juros, 'amort': amort,
        'nop': nop, 'outras_rec': outras_rec,
        'irpj_csll': irpj + csll,
        'resultado_original': resultado,
        'aj1': aj1, 'aj2': aj2, 'aj3': aj3, 'aj4': aj4, 'aj5': aj5,
        'resultado_corrigido': res_corr,
        'ebitda_corrigido': ebitda_corr,
        'rows_src': rows,
        'km': km,
    }

# ════════════════════════════════════════════════════════════
# PROCESSAMENTO — DRE DE OBRAS (múltiplas abas)
# ════════════════════════════════════════════════════════════
def processar_obras(wb, nome, ajustes, da_pct):
    EXCLUIR = {'obras em andamento','fluxo de caixa','rental','dashboard','fs energia-cnp'}
    abas = [s for s in wb.sheetnames if s.lower() not in EXCLUIR]
    resultado = {}

    for sname in abas:
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

        # Encontrar coluna Total
        header = rows[7] if len(rows) > 7 else []
        total_col = None
        for ci, v in enumerate(header):
            if str(v).strip().lower() == 'total':
                total_col = ci; break

        # Mapear linhas-chave
        km = {}
        for i, row in enumerate(rows, 1):
            b = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ''
            if b == 'faturamento total':       km['fat'] = i
            elif b == 'resultado operacional': km['res'] = i
            elif '200 - aquisição de imobilizado' in b: km['cap_imo'] = i
            elif '20702' in b:                km['cap_mov'] = i
            elif '20704' in b:                km['cap_maq'] = i
            elif '20706' in b:                km['cap_imov'] = i
            elif b == 'despesas financeiras':  km['fin'] = i
            elif '20407' in b or 'ir pessoa jurídica' in b: km['irpj'] = i
            elif '20412' in b or b == 'csll':  km['csll'] = i

        if 'fat' not in km or 'res' not in km:
            continue

        def gv(key):
            r = km.get(key)
            if r is None: return 0.0
            row = rows[r - 1]
            if total_col is not None:
                try: return float(row[total_col]) or 0.0
                except: return 0.0
            # fallback: soma dos meses
            total = 0.0
            for ci in range(2, len(row)):
                try: total += float(row[ci]) or 0.0
                except: pass
            return total

        fat  = gv('fat'); res = gv('res')
        capex = -(gv('cap_imo') + gv('cap_mov') + gv('cap_maq') + gv('cap_imov'))
        irpj_csll = -(gv('irpj') + gv('csll'))
        da = abs(fat) * da_pct

        aj1 = capex       if ajustes[0] else 0
        aj2 = irpj_csll   if ajustes[1] else 0
        aj4 = da          if ajustes[3] else 0
        res_corr = res + aj1 + aj2 + aj4

        resultado[sname] = {
            'faturamento': fat, 'resultado_original': res,
            'aj1': aj1, 'aj2': aj2, 'aj3': 0, 'aj4': aj4, 'aj5': 0,
            'resultado_corrigido': res_corr,
            'ebitda_corrigido': res_corr + abs(gv('fin')) + aj4,
            'margem_orig': res / fat if fat else 0,
            'margem_corr': res_corr / fat if fat else 0,
        }
    return resultado

# ════════════════════════════════════════════════════════════
# PROCESSAMENTO — DRE CONSOLIDADA ANUAL
# ════════════════════════════════════════════════════════════
def processar_consolidada(wb, nome, ajustes, da_pct):
    EXCLUIR = {'obras em andamento','fluxo de caixa','rental','dashboard'}
    abas = [s for s in wb.sheetnames if s.lower() not in EXCLUIR]
    if not abas: return None

    ws = wb[abas[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    km = {}
    for i, row in enumerate(rows, 1):
        code = ' '.join(str(row[0]).strip().split()) if row[0] else ''
        b    = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ''
        if code in ('01.01.01','01.01.01.01') and 'rec_bruta' not in km: km['rec_bruta'] = i
        elif code == '01.01.02' and 'deducoes' not in km:  km['deducoes'] = i
        elif code == '01.01.02.01.06' and 'irpj' not in km: km['irpj'] = i
        elif code == '02.01.02.05.04' and 'capex' not in km: km['capex'] = i
        elif code == '03.03' and 'desp_fin' not in km:    km['desp_fin'] = i
        elif code == '03.03.01.07' and 'amort' not in km: km['amort'] = i
        elif code in ('05',) and 'resultado' not in km:   km['resultado'] = i
        elif b in ('resultado operacional',) and 'resultado' not in km: km['resultado'] = i

    # Encontrar coluna total
    total_col = None
    for row in rows[:15]:
        for ci, v in enumerate(row):
            if str(v).strip().lower() in ('total','total geral'):
                total_col = ci; break
        if total_col is not None: break
    if total_col is None:
        for row in rows[5:20]:
            for ci in range(min(len(row)-1,40), 1, -1):
                try:
                    if row[ci] and abs(float(row[ci])) > 10000:
                        total_col = ci; break
                except: pass
            if total_col is not None: break
    if total_col is None: total_col = 2

    def gv(key):
        r = km.get(key)
        if r is None: return 0.0
        row = rows[r-1]
        try: return float(row[total_col]) or 0.0
        except: return 0.0

    rb   = gv('rec_bruta'); ded = gv('deducoes')
    irpj = gv('irpj'); capex = gv('capex')
    fin  = gv('desp_fin'); amort = gv('amort')
    res  = gv('resultado')
    da   = abs(rb) * da_pct

    aj1 = -capex if ajustes[0] else 0
    aj2 = -irpj  if ajustes[1] else 0
    aj3 = -amort if ajustes[2] else 0
    aj4 = da     if ajustes[3] else 0
    res_corr = res + aj1 + aj2 + aj3 + aj4
    fin_corr = fin - amort if ajustes[2] else fin
    ebitda   = res_corr - fin_corr + aj4

    return {
        'tipo': 'consolidada', 'arquivo': nome,
        'rec_bruta': rb, 'resultado_original': res,
        'capex_aj': aj1, 'impostos_aj': aj2, 'amort_aj': aj3, 'da_val': aj4,
        'resultado_corrigido': res_corr,
        'ebitda_corrigido': ebitda,
        'margem_ebitda_corr': ebitda / rb if rb else 0,
    }

# ════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL — MENSAL COMPLETO
# ════════════════════════════════════════════════════════════
def _excel_mensal(ws, d, ajustes, da_pct):
    """Gera aba DRE completa linha por linha para DRE mensal."""
    periodo = d['periodo']
    rows_src = d.get('rows_src', [])

    # Hierarquia
    def get_parent(code):
        p = code.rsplit('.', 1)
        return p[0] if len(p) > 1 else None

    todos_codes = set()
    for row in rows_src[1:]:
        code = ' '.join(str(row[0]).strip().split()) if row[0] else ''
        try:
            v = float(row[2])
            if code: todos_codes.add(code)
        except: pass

    codigos_mae = set()
    for code in todos_codes:
        if '.' not in code: continue  # raízes (01,02,03,04,05) são tratadas separadamente
        p = get_parent(code)
        if p and p in todos_codes and p != code:
            codigos_mae.add(p)

    H1 = {'01','02','03','04','05'}
    H2 = {'01.01','02.01','03.01','03.02','03.03','04.01'}

    AJUSTE_MAP = {
        '02.01.02.05.04': (0, '=-D{r}', 'Aj.1: CAPEX — compra de ativo vai para o Imobilizado, não é despesa'),
        '01.01.02.01.06': (1, '=-D{r}', 'Aj.2: IRPJ é imposto sobre o lucro — move para provisão ao final da DRE'),
        '01.01.02.01.05': (1, '=-D{r}', 'Aj.2: CSLL é imposto sobre o lucro — move para provisão ao final da DRE'),
        '03.03.01.07':    (2, '=-D{r}', 'Aj.3: Amortização de principal — reduz o Passivo no Balanço, não é despesa'),
        '01.01.03.01.04': (4, '=-D{r}', 'Aj.5: Empréstimos recebidos — são Passivo no Balanço, NÃO são receita'),
    }

    # Configuração de colunas
    ws.freeze_panes = 'C10'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 46
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 32

    # ── Cabeçalhos ──
    for r in range(1,6):
        for col in range(1,9): ws.cell(r,col).fill = fl('DARK')

    ws.merge_cells('A1:H1')
    c = ws['A1']; c.value = f'DRE CORRIGIDA — CAMPESATTO CONSTRUTORA LTDA  ·  {periodo}'
    c.font = ft(True,14,'WHITE'); c.fill = fl('DARK'); c.alignment = al('center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    c = ws['A2']; c.value = "CNPJ 03.722.632/0001-57  ·  Mirassol d'Oeste – MT"
    c.font = ft(False,9,'FFFFFF',True); c.fill = fl('DARK'); c.alignment = al('center')
    ws.row_dimensions[2].height = 14

    ws.merge_cells('A3:H3')
    c = ws['A3']
    c.value = ('🔵 AZUL = input Sienge   🟣 ROXO = ajuste (fórmula)   '
               '🟢 VERDE = corrigido (fórmula =SOMA filhas ou =D+E)   '
               '🟡 AMARELO = linha com ajuste aplicado')
    c.font = ft(False,9,'NOTE_FG',True); c.fill = fl('NOTE'); c.alignment = al('left',wrap=True)
    ws.row_dimensions[3].height = 18

    ws.merge_cells('A4:H4')
    c = ws['A4']
    c.value = ('Aj.1=CAPEX fora do custo  ·  Aj.2=IRPJ/CSLL para provisão  ·  '
               'Aj.3=Amortização para o Balanço  ·  Aj.4=D&A estimado  ·  '
               'Aj.5=Empréstimos recebidos removidos de Outras Receitas')
    c.font = ft(False,8,'FFFFFF',True); c.fill = fl('DARK'); c.alignment = al('left',wrap=True)
    ws.row_dimensions[4].height = 16

    ws.row_dimensions[5].height = 5
    ws.merge_cells('A6:H6')
    c = ws['A6']
    c.value = '  ⚙  PARÂMETRO — célula AZUL D7: altere o % do D&A → toda a DRE recalcula'
    c.font = ft(True,9,'WHITE'); c.fill = fl('MED'); c.alignment = al('left')
    ws.row_dimensions[6].height = 16

    ws.row_dimensions[7].height = 20
    for col in range(1,9): ws.cell(7,col).fill = fl('GRAY2')
    ws.cell(7,3,'D&A — % da Receita Bruta mensal (Ajuste 4):').font = ft(False,10,'333333')
    ws.cell(7,3).alignment = al('left')
    da_inp = ws.cell(7,4,da_pct)
    da_inp.font = ft(True,11,'0000FF'); da_inp.fill = fl('NOTE')
    da_inp.alignment = al('right'); da_inp.number_format = '0.00%'
    ws.cell(7,5,'← AZUL: altere → DRE recalcula').font = ft(False,9,'NOTE_FG',True)
    ws.cell(7,5).alignment = al('left')
    DA_REF = '$D$7'
    ws.row_dimensions[8].height = 4

    # Headers colunas
    ws.row_dimensions[9].height = 22
    hdrs = ['','','CÓDIGO / CONTA',f'ORIGINAL ({periodo})','AJUSTE','CORRIGIDO','% Rec Bruta','NOTA DO AJUSTE']
    hbgs = ['DARK','DARK','DARK','ORIG','ADJ','CORR','CORR','DARK']
    hfgs = ['WHITE','WHITE','WHITE','ORIG_FG','ADJ_FG','CORR_FG','CORR_FG','WHITE']
    for ci,(h,bg,fg) in enumerate(zip(hdrs,hbgs,hfgs),1):
        c = ws.cell(9,ci,h); c.font = ft(True,9,fg); c.fill = fl(bg)
        c.alignment = al('left' if ci<=3 else 'center'); c.border = bdr()
    ws.row_dimensions[10].height = 4

    # ── Linhas da DRE ──
    cur = 11
    ROW_MAP = {}
    RB_ROW = None; RES_ROW = None

    linhas = []
    for row in rows_src[1:]:
        code = ' '.join(str(row[0]).strip().split()) if row[0] else ''
        b    = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        try:
            val = float(row[2]) if row[2] and str(row[2]).strip() != periodo else None
        except: val = None
        if code and val is not None:
            linhas.append((code, b, val))

    for code, b, val in linhas:
        # Estilo
        if   code in H1:         bg,fg,bold,sz,rh = 'DARK','WHITE',True,11,22
        elif code in H2:         bg,fg,bold,sz,rh = 'MED','WHITE',True,10,20
        elif code in codigos_mae:bg,fg,bold,sz,rh = 'LIGHT','1A2744',True,10,19
        elif code in AJUSTE_MAP: bg,fg,bold,sz,rh = 'NOTE','333333',False,9,17
        else:
            lv = code.count('.')
            if lv <= 1:   bg,fg,bold,sz,rh = 'GRAY','333333',False,9,17
            elif lv == 2: bg,fg,bold,sz,rh = 'GRAY2','555555',False,9,16
            else:         bg,fg,bold,sz,rh = 'GRAY2','888888',False,8,15

        ws.row_dimensions[cur].height = rh
        for col in range(1,9): ws.cell(cur,col).fill = fl(bg)

        # Conta
        lv = code.count('.')
        indent = '  ' * min(lv, 5)
        ws.cell(cur,3,f"{indent}{code}  {b}").font = ft(bold,sz,fg)
        ws.cell(cur,3).alignment = al('left')

        # Col D — original (azul = input para folhas)
        dc = ws.cell(cur,4,val)
        dc.font = ft(bold,sz,'0000FF' if not bold else fg)
        dc.fill = fl('ORIG' if not bold else bg)
        dc.alignment = al('right'); dc.number_format = '#,##0;(#,##0);"-"'

        # Col E — ajuste
        has_adj = code in AJUSTE_MAP
        if has_adj:
            aj_idx, fml_tpl, nota = AJUSTE_MAP[code]
            if ajustes[aj_idx]:
                fml = fml_tpl.format(r=cur)
                ec = ws.cell(cur,5,fml)
                ec.font = ft(bold,sz,'ADJ_FG'); ec.fill = fl('ADJ')
                ec.alignment = al('right'); ec.number_format = '+#,##0;(#,##0);"-"'
                ws.cell(cur,8,nota).font = ft(False,8,'ADJ_FG',True)
                ws.cell(cur,8).alignment = al('left',wrap=True); ws.cell(cur,8).fill = fl('ADJ')
            else:
                has_adj = False

        if not has_adj:
            ws.cell(cur,5,'—').font = ft(False,8,'CCCCCC')
            ws.cell(cur,5).alignment = al('center')

        # Col F — corrigido (placeholder p/ mães, fórmula p/ folhas)
        fc = ws.cell(cur,6)
        if code in codigos_mae:
            fc.value = 'SOMA_DEPOIS'  # preenchido depois
        elif has_adj:
            fc.value = f'=D{cur}+E{cur}'
        else:
            fc.value = f'=D{cur}'
        fc.font = ft(bold,sz,'CORR_FG' if not bold else fg)
        fc.fill = fl('CORR' if not bold else bg)
        fc.alignment = al('right'); fc.number_format = '#,##0;(#,##0);"-"'

        # Col G — % Rec Bruta (grupos principais)
        if bold and RB_ROW:
            pc = ws.cell(cur,7,f'=IF($D${RB_ROW}<>0,F{cur}/$D${RB_ROW},0)')
            pc.font = ft(bold,sz,'555555'); pc.fill = fl(bg)
            pc.alignment = al('right'); pc.number_format = '0.0%;(0.0%);"-"'

        if bold:
            for col in range(1,9): ws.cell(cur,col).border = bdr(code in H1)

        ROW_MAP[code] = cur
        if code == '01.01.01': RB_ROW = cur
        if code == '05':       RES_ROW = cur
        cur += 1

    # ── Segunda passagem: fórmulas SOMA nas contas mãe ──
    for code in sorted(codigos_mae, key=lambda x: -x.count('.')):
        er = ROW_MAP.get(code)
        if not er: continue
        filhos = sorted([c for c in todos_codes if '.' in c and c.rsplit('.',1)[0] == code and c in ROW_MAP and c != code])
        if filhos:
            soma = '=' + '+'.join([f'F{ROW_MAP[f]}' for f in filhos])
        else:
            soma = f'=D{er}'
        bg = ('DARK' if code in H1 else 'MED' if code in H2 else 'LIGHT')
        bold = code in H1 | H2 | codigos_mae
        sz   = 11 if code in H1 else 10
        fc = ws.cell(er,6,soma)
        fc.font = ft(bold,sz,'CORR_FG'); fc.fill = fl('CORR' if code not in H1|H2 else bg)
        fc.alignment = al('right'); fc.number_format = '#,##0;(#,##0);"-"'

    # Conta 05: soma dos 4 grupos
    if RES_ROW:
        refs = [f'F{ROW_MAP[g]}' for g in ['01','02','03','04'] if g in ROW_MAP]
        ws.cell(RES_ROW,6,'='+'+'.join(refs)).font = ft(True,11,'CORR_FG')
        ws.cell(RES_ROW,6).fill = fl('DARK'); ws.cell(RES_ROW,6).alignment = al('right')
        ws.cell(RES_ROW,6).number_format = '#,##0;(#,##0);"-"'
        ws.cell(RES_ROW,4,'='+'+'.join([f'D{ROW_MAP[g]}' for g in ['01','02','03','04'] if g in ROW_MAP]))
        ws.cell(RES_ROW,4).font = ft(True,11,'WHITE'); ws.cell(RES_ROW,4).fill = fl('DARK')
        ws.cell(RES_ROW,4).alignment = al('right'); ws.cell(RES_ROW,4).number_format = '#,##0;(#,##0);"-"'

    # ── Bloco ajustes adicionais ──
    ws.row_dimensions[cur].height = 8; cur += 1
    ws.merge_cells(f'A{cur}:H{cur}')
    c = ws.cell(cur,1,'  ▼  AJUSTES ADICIONAIS (D&A e Provisão IRPJ)  ▼')
    c.font = ft(True,9,'WHITE'); c.fill = fl('MED'); c.alignment = al('center')
    ws.row_dimensions[cur].height = 16; cur += 1

    # D&A
    DA_ROW = cur; ws.row_dimensions[cur].height = 20
    for col in range(1,9): ws.cell(cur,col).fill = fl('ADJ')
    ws.cell(cur,3,'(+) D&A — Depreciação e Amortização add-back (Ajuste 4)').font = ft(True,10,'ADJ_FG')
    ws.cell(cur,3).alignment = al('left')
    ws.cell(cur,4,'—').font = ft(False,9,'CCCCCC'); ws.cell(cur,4).alignment = al('center')
    if RB_ROW and ajustes[3]:
        da_f = ws.cell(cur,5,f'=D{RB_ROW}*{DA_REF}')
    else:
        da_f = ws.cell(cur,5,0)
    da_f.font=ft(True,10,'ADJ_FG'); da_f.fill=fl('ADJ')
    da_f.alignment=al('right'); da_f.number_format='+#,##0;(#,##0);"-"'
    da_fc = ws.cell(cur,6,f'=E{cur}')
    da_fc.font=ft(True,10,'ADJ_FG'); da_fc.fill=fl('ADJ')
    da_fc.alignment=al('right'); da_fc.number_format='#,##0;(#,##0);"-"'
    ws.cell(cur,8,'Aj.4: D&A = Rec Bruta × % (célula D7). Altere para calibrar.').font=ft(False,8,'ADJ_FG',True)
    ws.cell(cur,8).alignment=al('left',wrap=True); ws.cell(cur,8).fill=fl('ADJ')
    for col in range(1,9): ws.cell(cur,col).border = bdr()
    cur += 1

    # Provisão IRPJ (informativa)
    PROV_ROW = cur; ws.row_dimensions[cur].height = 20
    for col in range(1,9): ws.cell(cur,col).fill = fl('ADJ')
    ws.cell(cur,3,'(-) Provisão IRPJ+CSLL — reposicionada aqui (informativo, efeito líquido = zero)').font=ft(True,10,'ADJ_FG')
    ws.cell(cur,3).alignment = al('left')
    ws.cell(cur,4,'—').font=ft(False,9,'CCCCCC'); ws.cell(cur,4).alignment=al('center')
    ir = ROW_MAP.get('01.01.02.01.06')
    if ir and ajustes[1]:
        pf = ws.cell(cur,5,f'=D{ir}')
        pf.font=ft(True,10,'ADJ_FG'); pf.fill=fl('ADJ')
        pf.alignment=al('right'); pf.number_format='#,##0;(#,##0);"-"'
    ws.cell(cur,8,'Aj.2: IRPJ saiu das deduções de receita (linha 01.01.02.01.06). Lançado aqui como provisão — efeito no resultado = zero.').font=ft(False,8,'ADJ_FG',True)
    ws.cell(cur,8).alignment=al('left',wrap=True); ws.cell(cur,8).fill=fl('ADJ')
    for col in range(1,9): ws.cell(cur,col).border = bdr()
    cur += 1
    ws.row_dimensions[cur].height = 8; cur += 1

    # ── Resultado Corrigido ──
    RES_CORR = cur; ws.row_dimensions[cur].height = 28
    for col in range(1,9): ws.cell(cur,col).fill = fl('DARK')
    ws.cell(cur,3,'★  RESULTADO OPERACIONAL CORRIGIDO').font=ft(True,13,'WHITE')
    ws.cell(cur,3).alignment = al('left')
    if RES_ROW:
        ws.cell(cur,4,f'=D{RES_ROW}').font=ft(True,12,'ORIG_FG')
        ws.cell(cur,4).fill=fl('ORIG'); ws.cell(cur,4).alignment=al('right')
        ws.cell(cur,4).number_format='#,##0;(#,##0);"-"'
        # Total ajustes = soma de todas as col E de linhas ajustadas + D&A
        aj_refs = [f'E{ROW_MAP[k]}' for k in AJUSTE_MAP if k in ROW_MAP
                   and ajustes[AJUSTE_MAP[k][0]]]
        aj_refs.append(f'E{DA_ROW}')
        aj_sum = '+'.join(aj_refs) if aj_refs else '0'
        ws.cell(cur,5,f'={aj_sum}').font=ft(True,12,'ADJ_FG')
        ws.cell(cur,5).fill=fl('ADJ'); ws.cell(cur,5).alignment=al('right')
        ws.cell(cur,5).number_format='+#,##0;(#,##0);"-"'
        # Corrigido = F(conta 05) + D&A apenas (Aj2 já se anula na árvore)
        ws.cell(cur,6,f'=F{RES_ROW}+E{DA_ROW}').font=ft(True,14,'CORR_FG')
        ws.cell(cur,6).fill=fl('CORR'); ws.cell(cur,6).alignment=al('right')
        ws.cell(cur,6).number_format='#,##0;(#,##0);"-"'
    for col in range(1,9): ws.cell(cur,col).border = bdr(True)
    cur += 1

    # ── EBITDA ──
    EBIT_ROW = cur; ws.row_dimensions[cur].height = 24
    for col in range(1,9): ws.cell(cur,col).fill = fl('TEAL')
    ws.cell(cur,3,'★  EBITDA Corrigido  (= Resultado Corrigido − Desp Financeiras Corrigidas + D&A)').font=ft(True,12,'WHITE')
    ws.cell(cur,3).alignment = al('left')
    fin_r = ROW_MAP.get('03.03')
    if RES_ROW and fin_r:
        ws.cell(cur,4,f'=D{RES_ROW}-D{fin_r}').font=ft(True,11,'999999')
        ws.cell(cur,4).fill=fl('TEAL'); ws.cell(cur,4).alignment=al('right')
        ws.cell(cur,4).number_format='#,##0;(#,##0);"-"'
        ws.cell(cur,6,f'=F{RES_CORR}-F{fin_r}+E{DA_ROW}').font=ft(True,14,'CORR_FG')
        ws.cell(cur,6).fill=fl('CORR'); ws.cell(cur,6).alignment=al('right')
        ws.cell(cur,6).number_format='#,##0;(#,##0);"-"'
    for col in range(1,9): ws.cell(cur,col).border = bdr(True)
    cur += 1

    # Margens
    for lbl, dr, fr in [
        ('  Margem EBITDA (% Rec Bruta)', EBIT_ROW, EBIT_ROW),
        ('  Margem Resultado Corrigido (% Rec Bruta)', RES_ROW, RES_CORR),
    ]:
        ws.row_dimensions[cur].height = 18
        for col in range(1,9): ws.cell(cur,col).fill = fl('CORR')
        ws.cell(cur,3,lbl).font=ft(False,10,'CORR_FG'); ws.cell(cur,3).alignment=al('left')
        if RB_ROW:
            ws.cell(cur,4,f'=IF($D${RB_ROW}<>0,D{dr}/$D${RB_ROW},0)').font=ft(True,10,'ORIG_FG')
            ws.cell(cur,4).fill=fl('CORR'); ws.cell(cur,4).alignment=al('right')
            ws.cell(cur,4).number_format='0.0%;(0.0%);"-"'
            ws.cell(cur,6,f'=IF($D${RB_ROW}<>0,F{fr}/$D${RB_ROW},0)').font=ft(True,11,'CORR_FG')
            ws.cell(cur,6).fill=fl('CORR'); ws.cell(cur,6).alignment=al('right')
            ws.cell(cur,6).number_format='0.0%;(0.0%);"-"'
        ws.cell(cur,5,'—').font=ft(False,9,'CCCCCC')
        ws.cell(cur,5).alignment=al('center'); ws.cell(cur,5).fill=fl('CORR')
        for col in range(1,9): ws.cell(cur,col).border = bdr()
        cur += 1

    return {'RB_ROW': RB_ROW, 'RES_ROW': RES_ROW, 'RES_CORR': RES_CORR,
            'EBIT_ROW': EBIT_ROW, 'DA_ROW': DA_ROW, 'ROW_MAP': ROW_MAP}

# ════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL — PAINEL RESUMO (obras / consolidada)
# ════════════════════════════════════════════════════════════
def _excel_resumo(ws, dados_lista):
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for col in ['C','D','E','F','G','H']: ws.column_dimensions[col].width = 17

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = f'PAINEL DRE CORRIGIDA — CAMPESATTO CONSTRUTORA  ·  {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    c.font = ft(True,12,'WHITE'); c.fill = fl('DARK'); c.alignment = al('center')
    ws.row_dimensions[1].height = 26

    for ci,h in enumerate(['','OBRA / PERÍODO','FATURAMENTO','RES. ORIGINAL','+CAPEX','+IRPJ/CSLL','+D&A','RES. CORRIGIDO'],1):
        c = ws.cell(4,ci,h)
        c.font = ft(True,9,'WHITE'); c.fill = fl('MED')
        c.alignment = al('center'); c.border = bdr()
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[2].height = 6; ws.row_dimensions[3].height = 6

    r = 5
    for tipo, d in dados_lista:
        if tipo == 'obras':
            for obra, od in d.items():
                bg = 'WHITE' if r%2==0 else 'GRAY2'
                for col in range(1,9): ws.cell(r,col).fill = fl(bg)
                ws.cell(r,2,obra).font=ft(False,9,'333333'); ws.cell(r,2).alignment=al('left')
                for ci,v in zip(range(3,9),[od['faturamento'],od['resultado_original'],
                                             od['aj1'],od['aj2'],od['aj4'],od['resultado_corrigido']]):
                    c = ws.cell(r,ci,v)
                    c.alignment = al('right'); c.number_format = '#,##0;(#,##0);"-"'
                    if ci==8:   c.font=ft(False,9,'1B5E20' if v>=0 else 'B71C1C')
                    elif ci in (5,6,7): c.font=ft(False,9,'4A148C')
                    else: c.font=ft(False,9,'333333')
                r+=1
        elif tipo == 'consolidada':
            for col in range(1,9): ws.cell(r,col).fill = fl('LIGHT')
            ws.cell(r,2,d['arquivo']).font=ft(True,10,'1A2744'); ws.cell(r,2).alignment=al('left')
            for ci,v in zip(range(3,9),[d['rec_bruta'],d['resultado_original'],
                                         d.get('capex_aj',0),d.get('impostos_aj',0),
                                         d.get('da_val',0),d['resultado_corrigido']]):
                c=ws.cell(r,ci,v); c.alignment=al('right')
                c.number_format='#,##0;(#,##0);"-"'
                c.font=ft(True,10,'1B5E20' if(ci==8 and v>=0) else '333333')
            r+=1

# ════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL — ORQUESTRADOR
# ════════════════════════════════════════════════════════════
def gerar_excel(dados_lista, ajustes, da_pct):
    wb = Workbook()
    first = True
    tem_obras_ou_consol = any(t in ('obras','consolidada') for t,_ in dados_lista)

    for tipo, d in dados_lista:
        if tipo == 'mensal':
            periodo_t = d['periodo'].replace('/','_')[:28]
            if first:
                ws = wb.active; ws.title = periodo_t; first = False
            else:
                ws = wb.create_sheet(periodo_t)
            _excel_mensal(ws, d, ajustes, da_pct)

            # Aba memória dos ajustes
            ws_mem = wb.create_sheet('Memória dos Ajustes')
            _excel_memoria(ws_mem, d, ws.title)

    if tem_obras_ou_consol:
        ws_p = wb.create_sheet('Painel Resumo') if not first else wb.active
        if first: ws_p.title = 'Painel Resumo'; first = False
        _excel_resumo(ws_p, [(t,d) for t,d in dados_lista if t in ('obras','consolidada')])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ════════════════════════════════════════════════════════════
# ABA MEMÓRIA DOS AJUSTES
# ════════════════════════════════════════════════════════════
def _excel_memoria(ws2, d, dre_tab_name):
    ws2.column_dimensions['A'].width = 4
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 44

    ws2.merge_cells('A1:F1')
    c = ws2['A1']
    c.value = f'MEMÓRIA DOS 5 AJUSTES — {d["periodo"]}'
    c.font = ft(True,12,'WHITE'); c.fill = fl('DARK'); c.alignment = al('center')
    ws2.row_dimensions[1].height = 26
    ws2.row_dimensions[2].height = 5

    for ci,h in enumerate(['','AJUSTE','CÓDIGO','DESCRIÇÃO','VALOR','EXPLICAÇÃO'],1):
        c = ws2.cell(3,ci,h)
        c.font=ft(True,9,'WHITE'); c.fill=fl('DARK')
        c.alignment=al('center'); ws2.row_dimensions[3].height=20

    DT = f"'{dre_tab_name}'"
    ROW_MAP = d.get('_row_map_for_memo', {})

    mem = [
        (1,'ORIG','CAPEX — Aquisição Imobilizado','02.01.02.05.04',
         'Compra de ativo vai para o Ativo Imobilizado, não é despesa de resultado.'),
        (2,'ADJ','IRPJ nas Deduções de Receita','01.01.02.01.06',
         'IRPJ é imposto sobre o LUCRO. O Sienge deduziu da receita bruta (incorreto). '
         'Reposicionado como provisão ao final da DRE. Efeito líquido no resultado = zero.'),
        (3,'ORIG','Amortização de Principal','03.03.01.07',
         'Pagamento do principal da dívida reduz o Passivo no Balanço. '
         'Não é despesa de resultado — apenas os JUROS ficam na DRE.'),
        (4,'ADJ','D&A Estimado','—',
         f'Depreciação estimada em {d["aj4"]/abs(d["rec_bruta"])*100:.1f}% da Receita Bruta. '
         'Solicite ao contador o razão de depreciação para o valor exato.'),
        (5,'NOTE','Empréstimos em Outras Receitas','01.01.03.01.04',
         'R$200k (Capital de Giro) + R$950k (Mútuo) lançados como Outras Receitas. '
         'Incorreto: recebimento de empréstimo cria um Passivo no Balanço, não é receita.'),
    ]

    vals_aj = [d['aj1'], d['aj2'], d['aj4'], d['aj4'], d['aj5']]

    for i,(num,bg_k,nome,code,explic) in enumerate(mem,1):
        r = 3+i; ws2.row_dimensions[r].height = 44
        for col in range(1,7): ws2.cell(r,col).fill = fl(bg_k)
        ws2.cell(r,2,f'Ajuste {num}: {nome}').font=ft(True,10,'333333')
        ws2.cell(r,2).alignment=al('left')
        ws2.cell(r,3,code).font=ft(False,9,'555555'); ws2.cell(r,3).alignment=al('center')
        ws2.cell(r,4,nome).font=ft(False,9,'333333'); ws2.cell(r,4).alignment=al('left')
        vc = ws2.cell(r,5,vals_aj[i-1])
        vc.font=ft(True,11,'4A148C'); vc.alignment=al('right')
        vc.number_format='+#,##0;(#,##0);"-"'
        ws2.cell(r,6,explic).font=ft(False,9,'444444',True)
        ws2.cell(r,6).alignment=al('left',wrap=True)
        for col in range(1,7): ws2.cell(r,col).border=bdr()

    # Totais
    r = 3+len(mem)+1; ws2.row_dimensions[r].height=22
    for col in range(1,7): ws2.cell(r,col).fill=fl('LIGHT')
    ws2.cell(r,2,'TOTAL DOS 5 AJUSTES').font=ft(True,11,'1A2744'); ws2.cell(r,2).alignment=al('left')
    total_aj = d['aj1']+d['aj2']+d['aj3']+d['aj4']+d['aj5']
    vc=ws2.cell(r,5,total_aj); vc.font=ft(True,12,'4A148C')
    vc.alignment=al('right'); vc.number_format='+#,##0;(#,##0);"-"'
    for col in range(1,7): ws2.cell(r,col).border=bdr(True)

    r+=2
    for lbl,v,bg_k,fg_k in [
        ('Resultado Original (Sienge)', d['resultado_original'], 'ORIG','ORIG_FG'),
        ('Total Ajustes', total_aj, 'ADJ','ADJ_FG'),
        ('★ Resultado Corrigido', d['resultado_corrigido'], 'CORR','CORR_FG'),
        ('★ EBITDA Corrigido', d['ebitda_corrigido'], 'CORR','CORR_FG'),
    ]:
        ws2.row_dimensions[r].height=22
        for col in range(1,7): ws2.cell(r,col).fill=fl(bg_k)
        ws2.cell(r,2,lbl).font=ft(True,11,fg_k); ws2.cell(r,2).alignment=al('left')
        vc=ws2.cell(r,5,v); vc.font=ft(True,13,fg_k)
        vc.alignment=al('right'); vc.number_format='#,##0;(#,##0);"-"'
        for col in range(1,7): ws2.cell(r,col).border=bdr(lbl.startswith('★'))
        r+=1

# ════════════════════════════════════════════════════════════
# INTERFACE STREAMLIT
# ════════════════════════════════════════════════════════════
def main():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1A2744,#2E4A7A);border-radius:12px;
    padding:22px 28px;margin-bottom:20px;display:flex;align-items:center;justify-content:space-between">
    <div>
    <div style="color:#fff;font-size:20px;font-weight:500">📊 DRE Sienge — Processador de Ajustes</div>
    <div style="color:rgba(255,255,255,0.55);font-size:13px;margin-top:4px">
    Campesatto Construtora Ltda  ·  5 ajustes automáticos  ·  Download DRE completa</div>
    </div>
    <div style="background:rgba(255,255,255,0.12);color:rgba(255,255,255,0.85);
    border-radius:20px;padding:5px 14px;font-size:12px">v4.0</div>
    </div>
    """, unsafe_allow_html=True)

    # ── SIDEBAR ──
    with st.sidebar:
        st.markdown("### ⚙️ Ajustes")
        st.markdown("---")
        aj1 = st.checkbox("Aj.1 — CAPEX fora do custo",    value=True,
                          help="Aquisição de Imobilizado vai para o Ativo, não é despesa")
        aj2 = st.checkbox("Aj.2 — IRPJ/CSLL para provisão", value=True,
                          help="Move IRPJ e CSLL para o final da DRE")
        aj3 = st.checkbox("Aj.3 — Amortização para Balanço", value=True,
                          help="Amortização de principal não é despesa de resultado")
        aj4 = st.checkbox("Aj.4 — D&A estimado",            value=True,
                          help="Adiciona estimativa de depreciação como despesa não-caixa")
        aj5 = st.checkbox("Aj.5 — Empréstimos em Receita",  value=True,
                          help="Remove empréstimos recebidos classificados como Outras Receitas")
        ajustes = [aj1, aj2, aj3, aj4, aj5]
        st.markdown("---")
        st.markdown("**D&A — % da Receita Bruta:**")
        da_pct = st.slider("", min_value=0.5, max_value=3.0, value=1.0, step=0.1, format="%.1f%%") / 100
        st.markdown("---")
        st.markdown("""<div style='font-size:11px;color:#6B7280;line-height:1.7'>
        <b>Cores no Excel gerado:</b><br>
        🔵 Azul = input Sienge<br>🟣 Roxo = ajuste<br>🟢 Verde = corrigido (fórmulas)<br>
        🟡 Amarelo = linha ajustada<br><br>
        <b>Contas mãe:</b> fórmula =SOMA(filhas)<br>
        <b>Altere D&A na célula D7</b> → tudo recalcula
        </div>""", unsafe_allow_html=True)

    # ── TABS ──
    tab_up, tab_res, tab_obras, tab_comp, tab_info = st.tabs([
        "📂 Upload", "📊 Resultado", "🏗️ Por Obra", "📈 Comparativo", "❓ Como funciona"
    ])

    # ── UPLOAD ──
    with tab_up:
        st.markdown("#### Faça upload dos arquivos exportados do Sienge")
        st.info("Aceita **.xlsx**, **.xlsm** — DRE mensal, obras, ou consolidada anual. Pode enviar vários de uma vez.")

        uploaded = st.file_uploader(
            "Selecione os arquivos",
            type=["xlsx","xlsm","xls"],
            accept_multiple_files=True,
            label_visibility="collapsed",
        )

        if uploaded:
            st.markdown(f"**{len(uploaded)} arquivo(s):**")
            for f in uploaded:
                st.markdown(f"📄 `{f.name}` — {f.size//1024} KB")

            st.markdown("")
            if st.button("⚡ Processar e aplicar ajustes"):
                dados = []
                prog = st.progress(0, "Iniciando...")
                for idx, f in enumerate(uploaded):
                    prog.progress(int(idx/len(uploaded)*85), f"Processando {f.name}...")
                    try:
                        raw = io.BytesIO(f.read())
                        wb = load_workbook(raw, keep_vba=False, data_only=True)
                        tipo = detectar_tipo(wb)
                        if tipo == 'mensal':
                            r = processar_mensal(wb, f.name, ajustes, da_pct)
                            dados.append(('mensal', r))
                        elif tipo == 'obras':
                            r = processar_obras(wb, f.name, ajustes, da_pct)
                            if r: dados.append(('obras', r))
                        else:
                            r = processar_consolidada(wb, f.name, ajustes, da_pct)
                            if r: dados.append(('consolidada', r))
                    except Exception as e:
                        st.warning(f"⚠️ Erro em {f.name}: {e}")

                prog.progress(100, "Concluído!")
                st.session_state['dados'] = dados
                st.session_state['ajustes'] = ajustes
                st.session_state['da_pct'] = da_pct
                if dados:
                    st.success(f"✅ {len(dados)} arquivo(s) processado(s). Acesse as abas acima.")
                else:
                    st.error("Nenhum arquivo reconhecido. Verifique se são DREs exportadas do Sienge.")

    # ── RESULTADO ──
    with tab_res:
        dados = st.session_state.get('dados', [])
        if not dados:
            st.info("Faça o upload e processe os arquivos na aba **Upload** primeiro.")
        else:
            ajustes_us = st.session_state['ajustes']
            da_pct_us  = st.session_state['da_pct']

            # Totais
            tot_fat=tot_orig=tot_corr=tot_ebitda = 0
            for tipo, d in dados:
                if tipo == 'mensal':
                    tot_fat   += abs(d['rec_bruta'])
                    tot_orig  += d['resultado_original']
                    tot_corr  += d['resultado_corrigido']
                    tot_ebitda+= d['ebitda_corrigido']
                elif tipo == 'consolidada':
                    tot_fat   += abs(d['rec_bruta'])
                    tot_orig  += d['resultado_original']
                    tot_corr  += d['resultado_corrigido']
                    tot_ebitda+= d['ebitda_corrigido']
                elif tipo == 'obras':
                    for od in d.values():
                        tot_fat   += abs(od['faturamento'])
                        tot_orig  += od['resultado_original']
                        tot_corr  += od['resultado_corrigido']

            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Receita/Faturamento", fmtM(tot_fat))
            delta_orig = f"{'🔴' if tot_orig<0 else '🟢'} DRE Sienge"
            c2.metric("Resultado Original", fmtM(tot_orig), delta=delta_orig, delta_color="off")
            delta_corr = f"+{fmtM(tot_corr-tot_orig)} pelos ajustes"
            c3.metric("Resultado Corrigido", fmtM(tot_corr), delta=delta_corr,
                      delta_color="normal" if tot_corr > tot_orig else "inverse")
            c4.metric("EBITDA Corrigido", fmtM(tot_ebitda),
                      delta=pct_str(tot_ebitda, tot_fat) + " da RB")

            # Gráfico cascata
            if tot_orig != 0:
                st.markdown("---")
                st.markdown("#### Cascata de ajustes")
                tot_aj = [0]*5
                for tipo, d in dados:
                    if tipo == 'mensal':
                        for i, k in enumerate(['aj1','aj2','aj3','aj4','aj5']):
                            tot_aj[i] += d.get(k, 0)
                    elif tipo == 'consolidada':
                        tot_aj[0] += d.get('capex_aj',0)
                        tot_aj[1] += d.get('impostos_aj',0)
                        tot_aj[2] += d.get('amort_aj',0)
                        tot_aj[3] += d.get('da_val',0)

                fig = go.Figure(go.Waterfall(
                    orientation="v",
                    measure=["absolute","relative","relative","relative","relative","relative","total"],
                    x=["Original","Aj.1 CAPEX","Aj.2 IRPJ","Aj.3 Amort","Aj.4 D&A","Aj.5 Emprést","Corrigido"],
                    y=[tot_orig]+tot_aj+[tot_corr],
                    connector={"line":{"color":"#E2E8F0","width":1}},
                    decreasing={"marker":{"color":"#DC2626"}},
                    increasing={"marker":{"color":"#059669"}},
                    totals={"marker":{"color":"#1A2744"}},
                    text=[fmtM(v) for v in [tot_orig]+tot_aj+[tot_corr]],
                    textposition="outside",
                ))
                fig.update_layout(height=350, margin=dict(l=20,r=20,t=20,b=40),
                                  paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                  showlegend=False,
                                  yaxis=dict(gridcolor="#F1F5F9",zeroline=True,zerolinecolor="#CBD5E1"),
                                  font=dict(family="Arial",size=11))
                st.plotly_chart(fig, use_container_width=True)

            # DRE mensal resumida
            for tipo, d in dados:
                if tipo == 'mensal':
                    st.markdown("---")
                    st.markdown(f"#### DRE Mensal — {d['periodo']}  ·  `{d['arquivo']}`")
                    linhas = [
                        ("RECEITA BRUTA", d['rec_bruta'], 0, d['rec_bruta']),
                        ("(-) Deduções", d['deducoes'], d['aj2'], d['deducoes']-d['aj2']),
                        ("  ↳ Aj.2: IRPJ reposicionado", d['irpj_csll'], -d['irpj_csll'], 0),
                        ("Outras Receitas Op.", d['outras_rec'], d['aj5'], d['outras_rec']+d['aj5']),
                        ("  ↳ Aj.5: Empréstimos removidos", d['outras_rec']-0.147, d['aj5'], 0),
                        ("(-) Custos Diretos", d['custos_dir'], d['aj1'], d['custos_dir']+d['aj1']),
                        ("  ↳ Aj.1: CAPEX retirado", d['capex'], -d['capex'], 0),
                        ("(-) Custos Indiretos", d['custos_ind'], 0, d['custos_ind']),
                        ("(-) Desp Administrativas", d['desp_adm'], 0, d['desp_adm']),
                        ("(-) Desp Tributárias", d['desp_trib'], 0, d['desp_trib']),
                        ("(-) Desp Financeiras", d['desp_fin'], d['aj3'], d['desp_fin']+d['aj3']),
                        ("  ↳ Aj.3: Amortização retirada", d['amort'], -d['amort'], 0),
                        ("(+/-) Outras NOP", d['nop'], 0, d['nop']),
                        ("(+) D&A add-back (Aj.4)", 0, d['aj4'], d['aj4']),
                        ("★ RESULTADO CORRIGIDO", d['resultado_original'],
                         d['aj1']+d['aj2']+d['aj3']+d['aj4']+d['aj5'], d['resultado_corrigido']),
                    ]
                    rows_html = ""
                    for lbl,orig,adj,corr in linhas:
                        is_tot = lbl.startswith('★')
                        is_sub = lbl.startswith('  ↳')
                        style = ("background:#E8F5E9;font-weight:600" if is_tot
                                 else "background:#EDE7F6;font-size:12px" if is_sub
                                 else "")
                        def fv(v):
                            if isinstance(v, str): return v
                            s = fmtK(v)
                            return f"<span style='color:{'#059669' if float(v)>=0 else '#DC2626'}'>{s}</span>"
                        rows_html += f"""<tr style="{style}">
                            <td style="padding:6px 10px;font-size:13px">{lbl}</td>
                            <td style="text-align:right;padding:6px 10px;font-size:12px;color:#E65100">{fv(orig)}</td>
                            <td style="text-align:right;padding:6px 10px;font-size:12px;color:#4A148C">{fv(adj)}</td>
                            <td style="text-align:right;padding:6px 10px;font-size:13px;font-weight:{'600' if is_tot else '400'}">{fv(corr)}</td>
                        </tr>"""
                    st.markdown(f"""<table style="width:100%;border-collapse:collapse;border:1px solid #E2E8F0;border-radius:8px;overflow:hidden">
                    <thead><tr style="background:#1A2744;color:#fff">
                    <th style="padding:8px 10px;text-align:left;font-size:12px">Linha</th>
                    <th style="padding:8px 10px;text-align:right;font-size:12px">Original</th>
                    <th style="padding:8px 10px;text-align:right;font-size:12px">Ajuste</th>
                    <th style="padding:8px 10px;text-align:right;font-size:12px">Corrigido</th>
                    </tr></thead><tbody>{rows_html}</tbody></table>""", unsafe_allow_html=True)

            st.markdown("---")
            # Download
            try:
                excel_bytes = gerar_excel(dados, ajustes_us, da_pct_us)
                st.download_button(
                    label="⬇️ Baixar DRE Corrigida — Excel completo (.xlsx)",
                    data=excel_bytes,
                    file_name=f"DRE_Corrigida_Campesatto_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Erro ao gerar Excel: {e}")

    # ── POR OBRA ──
    with tab_obras:
        dados = st.session_state.get('dados', [])
        obras_dict = {}
        for tipo, d in dados:
            if tipo == 'obras': obras_dict.update(d)

        if not obras_dict:
            st.info("Nenhuma DRE de obra encontrada. Faça upload do arquivo de resultados de obras.")
        else:
            obra_sel = st.selectbox("Selecione a obra:", list(obras_dict.keys()))
            if obra_sel:
                od = obras_dict[obra_sel]
                c1,c2,c3 = st.columns(3)
                c1.metric("Faturamento", fmtM(od['faturamento']))
                c2.metric("Resultado Original", fmtM(od['resultado_original']))
                c3.metric("Resultado Corrigido", fmtM(od['resultado_corrigido']),
                          delta=pct_str(od['resultado_corrigido'], od['faturamento'])+" margem")

    # ── COMPARATIVO ──
    with tab_comp:
        dados = st.session_state.get('dados', [])
        mensais = [(t,d) for t,d in dados if t=='mensal']
        consols = [(t,d) for t,d in dados if t=='consolidada']

        if not mensais and not consols:
            st.info("Processe os arquivos primeiro.")
        else:
            if consols:
                st.markdown("#### Comparativo: Sienge × Corrigida × SPED/Balanço 2024")
                rows_cmp = ""
                for _,d in consols:
                    rows_cmp += f"<tr style='background:#F5F7FA'><td colspan='4' style='padding:6px 10px;font-weight:600;font-size:13px'>{d['arquivo']}</td></tr>"
                    for lbl,orig_k,corr_k,sped in [
                        ("Receita Bruta","rec_bruta","rec_bruta",90_349_135),
                        ("Resultado","resultado_original","resultado_corrigido",3_993_382),
                        ("EBITDA","resultado_original","ebitda_corrigido",12_363_432),
                    ]:
                        o=d.get(orig_k,0); c2v=d.get(corr_k,0)
                        def fvc(v):
                            s=fmtK(v)
                            return f"<span style='color:{'#059669' if v>=0 else '#DC2626'}'>{s}</span>"
                        rows_cmp += f"""<tr>
                            <td style='padding:6px 10px;font-size:13px'>{lbl}</td>
                            <td style='text-align:right;padding:6px 10px;color:#E65100;font-size:12px'>{fvc(o)}</td>
                            <td style='text-align:right;padding:6px 10px;color:#4A148C;font-size:12px'>{fvc(c2v)}</td>
                            <td style='text-align:right;padding:6px 10px;color:#059669;font-size:12px'>{fvc(sped)}</td>
                        </tr>"""
                st.markdown(f"""<table style="width:100%;border-collapse:collapse;border:1px solid #E2E8F0">
                <thead><tr style="background:#1A2744;color:#fff">
                <th style="padding:8px 10px;text-align:left;font-size:12px">Indicador</th>
                <th style="padding:8px 10px;text-align:right;font-size:12px">🟠 Sienge</th>
                <th style="padding:8px 10px;text-align:right;font-size:12px">🟣 Corrigida</th>
                <th style="padding:8px 10px;text-align:right;font-size:12px">🟢 SPED/Balanço 2024</th>
                </tr></thead><tbody>{rows_cmp}</tbody></table>""", unsafe_allow_html=True)

    # ── COMO FUNCIONA ──
    with tab_info:
        st.markdown("#### Por que a DRE do Sienge difere do Balanço?")
        st.markdown("""
O Sienge é ótimo para controle operacional, mas mistura eventos de resultado com eventos de caixa e patrimônio.

| Distorção | Sienge lança como | Correto |
|---|---|---|
| Compra de máquina (CAPEX) | Custo dos serviços | Ativo Imobilizado |
| Amortização de dívida | Despesa financeira | Redução de Passivo |
| Empréstimo recebido | Outras Receitas | Passivo (dívida nova) |
| IRPJ/CSLL | Dedução de receita | Provisão no final da DRE |
| Depreciação (D&A) | Não segregado | Despesa não-caixa |

**Diferença confirmada em 2024:** DRE Sienge = -R$34,9M prejuízo. DRE SPED/Balanço = +R$3,99M lucro. Diferença de R$38,9M explicada pelos 4 ajustes principais.
        """)
        st.warning("⚠️ Esta DRE corrigida é **gerencial**. Para apresentar ao banco, use sempre o documento do SPED/Balanço elaborado pelo contador (Luiz Carlos Zimermann, CRC MT-003808/O-9).")

if __name__ == "__main__":
    main()

                
